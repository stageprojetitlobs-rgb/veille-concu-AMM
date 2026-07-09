"""Source AMM Vietnam — Cục Thú y (Department of Animal Health).

Le Cục Thú y publie la « Danh mục thuốc thú y được phép lưu hành tại Việt Nam »
(liste des médicaments vétérinaires autorisés) sous forme d'un **fichier Excel
public** sur cucthuy.gov.vn. Document de consultation publique → conforme.

Le classeur a 3 onglets utiles : médicaments fabriqués localement, **importés**
(le plus pertinent pour Lobs : concurrents étrangers), et aquaculture. Le nom de
la société figure en ligne-titre groupant ses produits (report en cascade, comme
l'ONSSA Maroc). Colonnes : TT | Tên thuốc | Hoạt chất chính | Dạng đóng gói |
Khối lượng | Công dụng | Số đăng ký.

Conformité : fichier public officiel, un GET par run, aucune donnée personnelle
(sociétés = personnes morales).
"""
from __future__ import annotations

import logging
import re
import tempfile

import httpx
import openpyxl

from veille.schema import Record, RecordType
from veille.sources.base import Source

log = logging.getLogger(__name__)

_COMPANY_KW = ("công ty", "cong ty", "cty", "company", "doanh nghiệp")


def _txt(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def _is_company(name_cell: str, tt_cell: str) -> bool:
    # Ligne-titre société : pas de n° d'ordre, et libellé type « 1. CÔNG TY … ».
    if tt_cell and tt_cell.strip().isdigit():
        return False
    low = name_cell.lower()
    return any(k in low for k in _COMPANY_KW)


def _clean_company(name: str) -> str:
    return re.sub(r"^\d+[\.\)]\s*", "", name).strip()


def _molecules(hoat_chat: str) -> list[str]:
    if not hoat_chat:
        return []
    parts = re.split(r"[,;/]+", hoat_chat)
    return [p.strip() for p in parts if p.strip()]


class CucthuyVietnamSource(Source):
    """Liste Cục Thú y (Vietnam). Config attendue (config.yaml) :

        sources:
          cucthuy_vietnam:
            enabled: true
            url_xlsx: "https://cucthuy.gov.vn/documents/.../...xlsx/...?t=..."
            sheets: ["PL1A. Thuoc thu y san xuat", "PL1B. Thuoc thu y nhap khau",
                     "PL1C. Thuoc thu y thuy san"]
            inclure_tous_produits: true
    """

    name = "cucthuy_vietnam"

    def fetch(self) -> list[Record]:
        url = self.cfg.get("url_xlsx")
        if not url:
            log.warning("cucthuy_vietnam : url_xlsx absente de la config")
            return []

        try:
            path = self._download(url)
        except httpx.HTTPError as exc:
            log.error("cucthuy_vietnam : échec téléchargement (%s)", exc)
            return []

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            log.error("cucthuy_vietnam : Excel illisible (%s)", exc)
            return []

        sheets = self.cfg.get("sheets") or wb.sheetnames
        inclure_tous = self.cfg.get("inclure_tous_produits", True)
        records: list[Record] = []
        seen: set[str] = set()

        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            cols = {}          # nom logique -> index colonne (détecté à l'en-tête)
            societe = ""
            for row in ws.iter_rows(values_only=True):
                cells = [_txt(c) for c in row]
                if not any(cells):
                    continue

                # En-tête de colonnes (contient « Tên thuốc »).
                joined = " ".join(cells).lower()
                if "tên thuốc" in joined or "ten thuoc" in joined:
                    for i, c in enumerate(cells):
                        cl = c.lower()
                        if "tên thuốc" in cl or "ten thuoc" in cl:
                            cols["produit"] = i
                        elif "hoạt chất" in cl or "hoat chat" in cl:
                            cols["molecules"] = i
                        elif "công dụng" in cl or "cong dung" in cl:
                            cols["usage"] = i
                        elif "số đăng ký" in cl or "so dang ky" in cl or "đăng ký" in cl:
                            cols["reg_no"] = i
                    continue
                if "produit" not in cols:
                    continue

                tt = cells[0] if cells else ""
                name_cell = cells[cols["produit"]] if cols["produit"] < len(cells) else ""

                # Ligne-titre société → mise à jour du contexte.
                if _is_company(" ".join(cells), tt):
                    societe = _clean_company(next((c for c in cells if c), ""))
                    continue
                if not name_cell:
                    continue
                # Ligne de données : doit avoir un n° d'ordre.
                if not tt or not tt.split(".")[0].strip().isdigit():
                    continue

                reg_no = cells[cols["reg_no"]] if cols.get("reg_no", -1) < len(cells) and "reg_no" in cols else ""
                uid = f"VN|{reg_no or societe}|{name_cell}".lower()
                if uid in seen:
                    continue
                seen.add(uid)

                concurrent = self.settings.matched_concurrent(societe) if societe else None
                if not concurrent and not inclure_tous:
                    continue

                hoat = cells[cols["molecules"]] if "molecules" in cols and cols["molecules"] < len(cells) else ""
                usage = cells[cols["usage"]] if "usage" in cols and cols["usage"] < len(cells) else ""
                tags = self.settings.keywords_in(f"{name_cell} {hoat} {usage}")

                rec = Record(
                    source=self.name,
                    source_uid=uid,
                    record_type=RecordType.NOUVELLE_AMM,
                    concurrent=concurrent,
                    produit=name_cell,
                    molecules=_molecules(hoat),
                    pays="VN",
                    url=url,
                    date_source=None,
                    tags=tags,
                    extra={
                        "titulaire": societe,
                        "usage": usage[:300],
                        "reg_no": reg_no,
                        "onglet": sheet_name,
                    },
                )
                rec.compute_hashes()
                records.append(rec)

        wb.close()
        log.info("cucthuy_vietnam : %d enregistrement(s) retenu(s)", len(records))
        return records

    def _download(self, url: str) -> str:
        with httpx.Client(
            headers={"User-Agent": self.settings.user_agent,
                     "Referer": "https://cucthuy.gov.vn/en/danh-muc-thuoc-thu-y"},
            timeout=max(self.settings.http_timeout_s, 90),
            follow_redirects=True,
            verify=False,
        ) as c:
            # Le portail Liferay sert parfois un corps vide sans session préalable.
            try:
                c.get("https://cucthuy.gov.vn/en/danh-muc-thuoc-thu-y")
            except httpx.HTTPError:
                pass
            resp = c.get(url)
            resp.raise_for_status()
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(resp.content)
        tmp.close()
        return tmp.name
